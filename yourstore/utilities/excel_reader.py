import openpyxl


def get_data_from_excel2():
    workbook=openpyxl.load_workbook("C:/Users/Vinuthreddy/Documents/Book6.xlsx")
    sheet=workbook["Sheet1"]

    values=[]
    for row in range(1,sheet.max_row+1):
        nested_values=[]
        for column in range(1,sheet.max_column+1):
            data=sheet.cell(row,column)
            nested_values.append(data.value)
        values.append(nested_values)


    return values


def get_data_from_excel():
    workbook=openpyxl.load_workbook("C:/Users/Vinuthreddy/Documents/Book5.xlsx")
    sheet=workbook["Sheet1"]

    values=[]
    for row in range(1,sheet.max_row+1):
        nested_values=[]
        for column in range(1,sheet.max_column+1):
            data=sheet.cell(row,column)
            nested_values.append(data.value)
        values.append(nested_values)


    return values